"""Historial de lecturas de un Contrato: consolida las Lecturas que
realmente se usaron para generar una Factura ya emitida —no las capturadas
pero todavía sin facturar, ni las que hayan quedado fuera de cualquier
periodo facturado— con el consumo calculado respecto a la lectura anterior
usando la misma lógica que el motor de facturación (`core.calculo_consumo`).

Usado por la vista de solo lectura en `core/admin.py` (botón "Historial de
lecturas" en Contrato) y por `generar_excel_historial` para su descarga.
"""
from dataclasses import dataclass
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from core.calculo_consumo import calcular_consumo
from core.models import Asignacion, Contrato, Factura, Lectura

AZUL_OSCURO = "1F3D5C"


@dataclass
class FilaHistorial:
    equipo_numero_serie: str
    equipo_marca_modelo: str
    fecha: date
    lectura_bn: int
    lectura_color: int
    consumo_bn: int
    consumo_color: int
    es_primera_lectura: bool
    origen: str
    estado_auditoria: str
    factura_periodo: str


def historial_de_contrato(contrato: Contrato) -> list[FilaHistorial]:
    """Lecturas de todas las Asignaciones que alguna vez tuvo este contrato
    (no solo las activas) que caen dentro del rango fecha_inicio/fecha_fin de
    alguna Factura ya emitida, ordenadas por equipo y fecha."""
    facturas = list(
        Factura.objects.filter(
            contrato=contrato, fecha_inicio__isnull=False, fecha_fin__isnull=False
        ).order_by("fecha_inicio")
    )
    if not facturas:
        return []

    asignaciones = Asignacion.objects.filter(contrato=contrato).select_related("equipo")

    filas = []
    for asignacion in asignaciones:
        lecturas = Lectura.objects.filter(asignacion=asignacion).order_by("fecha")
        for lectura in lecturas:
            factura = _factura_que_incluye(facturas, lectura.fecha)
            if factura is None:
                continue  # capturada pero todavía no facturada, o fuera de todo periodo facturado

            resultado_consumo = calcular_consumo(asignacion, lectura)
            filas.append(
                FilaHistorial(
                    equipo_numero_serie=asignacion.equipo.numero_serie,
                    equipo_marca_modelo=f"{asignacion.equipo.marca} {asignacion.equipo.modelo}",
                    fecha=lectura.fecha,
                    lectura_bn=lectura.lectura_bn,
                    lectura_color=lectura.lectura_color,
                    consumo_bn=resultado_consumo.consumo_bn,
                    consumo_color=resultado_consumo.consumo_color,
                    es_primera_lectura=resultado_consumo.es_primera_lectura_asignacion,
                    origen=lectura.origen,
                    estado_auditoria=lectura.estado_auditoria,
                    factura_periodo=f"{factura.periodo_mes:02d}/{factura.periodo_anio}",
                )
            )

    filas.sort(key=lambda f: (f.equipo_numero_serie, f.fecha))
    return filas


def _factura_que_incluye(facturas: list[Factura], fecha: date) -> Factura | None:
    for factura in facturas:
        if factura.fecha_inicio <= fecha <= factura.fecha_fin:
            return factura
    return None


def generar_excel_historial(contrato: Contrato, filas: list[FilaHistorial]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial"

    ws["A1"] = f"Historial de lecturas facturadas — Contrato {contrato.numero_contrato}"
    ws["A1"].font = Font(size=14, bold=True, color=AZUL_OSCURO)
    ws.merge_cells("A1:J1")

    encabezados = [
        "Equipo",
        "Marca / Modelo",
        "Fecha",
        "Lectura BN",
        "Lectura Color",
        "Consumo BN",
        "Consumo Color",
        "Nota",
        "Origen",
        "Auditoría",
        "Factura (periodo)",
    ]
    fila_encabezado = 3
    fuente_encabezado = Font(bold=True, color="FFFFFF")
    relleno_encabezado = PatternFill("solid", fgColor=AZUL_OSCURO)
    for col, texto in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila_encabezado, column=col, value=texto)
        celda.font = fuente_encabezado
        celda.fill = relleno_encabezado

    fila = fila_encabezado + 1
    for f in filas:
        ws.cell(row=fila, column=1, value=f.equipo_numero_serie)
        ws.cell(row=fila, column=2, value=f.equipo_marca_modelo)
        ws.cell(row=fila, column=3, value=f.fecha.strftime("%d/%m/%Y"))
        ws.cell(row=fila, column=4, value=f.lectura_bn)
        ws.cell(row=fila, column=5, value=f.lectura_color)
        ws.cell(row=fila, column=6, value=f.consumo_bn)
        ws.cell(row=fila, column=7, value=f.consumo_color)
        ws.cell(row=fila, column=8, value="Contra lectura de referencia inicial" if f.es_primera_lectura else "")
        ws.cell(row=fila, column=9, value=f.origen)
        ws.cell(row=fila, column=10, value=f.estado_auditoria)
        ws.cell(row=fila, column=11, value=f.factura_periodo)
        fila += 1

    if not filas:
        ws.cell(row=fila, column=1, value="Este contrato no tiene lecturas facturadas todavía.")

    anchos = [16, 26, 12, 12, 14, 12, 14, 32, 12, 12, 16]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
