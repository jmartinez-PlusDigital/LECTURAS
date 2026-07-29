from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ImagenXL
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.models import Contrato, Equipo
from facturacion import ResultadoFacturacion

from .pdf import MESES_ES, _tarifa

AZUL_OSCURO = "1F3D5C"
GRIS_CLARO = "F5F7F9"
GRIS_BORDE = "D9DEE4"
FORMATO_MONEDA = {"MXN": '"$"#,##0.00', "USD": '"US$"#,##0.00'}
FORMATO_ENTERO = "#,##0"
ALTO_LOGO_PX = 64


def _insertar_logo(ws, emisor, celda: str) -> None:
    """Inserta el logo del emisor en `celda`, escalado a una altura fija
    manteniendo su proporción. No hace nada si el emisor no tiene logo
    cargado todavía (ver core.models.EmpresaEmisora)."""
    if emisor is None or not emisor.logo:
        return
    try:
        with emisor.logo.open("rb") as archivo:
            datos = BytesIO(archivo.read())
    except (FileNotFoundError, ValueError):
        return
    imagen = ImagenXL(datos)
    proporcion = imagen.width / imagen.height
    imagen.height = ALTO_LOGO_PX
    imagen.width = ALTO_LOGO_PX * proporcion
    ws.add_image(imagen, celda)


def generar_excel_factura(resultado: ResultadoFacturacion, contrato: Contrato) -> bytes:
    """Genera el Excel de desglose de consumo y facturación de un contrato/periodo."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Factura"

    fuente_titulo = Font(size=14, bold=True, color=AZUL_OSCURO)
    fuente_encabezado_tabla = Font(bold=True, color="FFFFFF")
    relleno_encabezado_tabla = PatternFill("solid", fgColor=AZUL_OSCURO)
    fuente_etiqueta = Font(bold=True)
    relleno_alterno = PatternFill("solid", fgColor=GRIS_CLARO)
    borde_fino = Side(style="thin", color=GRIS_BORDE)
    borde_celda = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)

    ws.row_dimensions[1].height = 50
    ws["A1"] = f"Contrato {contrato.numero_contrato} — {MESES_ES[resultado.periodo_mes]} {resultado.periodo_anio}"
    ws["A1"].font = fuente_titulo
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells("A1:D1")
    _insertar_logo(ws, resultado.emisor, "F1")

    ws["A3"] = "Cliente"
    ws["A3"].font = fuente_etiqueta
    ws["B3"] = contrato.cliente.nombre
    ws["A4"] = "Periodo"
    ws["A4"].font = fuente_etiqueta
    ws["B4"] = f"{resultado.fecha_inicio.strftime('%d/%m/%Y')} - {resultado.fecha_fin.strftime('%d/%m/%Y')}"
    ws["A5"] = "Moneda"
    ws["A5"].font = fuente_etiqueta
    ws["B5"] = resultado.moneda

    formato_moneda = FORMATO_MONEDA.get(resultado.moneda, '#,##0.00 "' + resultado.moneda + '"')
    fila_encabezado_tabla = 7
    encabezados = [
        "Número de serie",
        "Marca / Modelo",
        "Fecha lect. anterior",
        "Lect. anterior BN",
        "Lect. anterior Color",
        "Fecha lect. actual",
        "Lect. actual BN",
        "Lect. actual Color",
        "Consumo BN",
        "Consumo Color",
    ]
    ws.row_dimensions[fila_encabezado_tabla].height = 32
    for col, texto in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila_encabezado_tabla, column=col, value=texto)
        celda.font = fuente_encabezado_tabla
        celda.fill = relleno_encabezado_tabla
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = borde_celda

    series = [c.equipo_numero_serie for c in resultado.consumo_por_equipo]
    marcas_modelos = {
        e.numero_serie: f"{e.marca} {e.modelo}"
        for e in Equipo.objects.filter(numero_serie__in=series).only("numero_serie", "marca", "modelo")
    }

    columnas_numericas = {4, 5, 7, 8, 9, 10}
    fila = fila_encabezado_tabla + 1
    for consumo in resultado.consumo_por_equipo:
        valores = (
            consumo.equipo_numero_serie,
            marcas_modelos.get(consumo.equipo_numero_serie, ""),
            consumo.fecha_lectura_anterior.strftime("%d/%m/%Y"),
            consumo.lectura_anterior_bn,
            consumo.lectura_anterior_color,
            consumo.fecha_lectura_actual.strftime("%d/%m/%Y"),
            consumo.lectura_actual_bn,
            consumo.lectura_actual_color,
            consumo.consumo_bn,
            consumo.consumo_color,
        )
        for col, valor in enumerate(valores, start=1):
            celda = ws.cell(row=fila, column=col, value=valor)
            celda.border = borde_celda
            if col in columnas_numericas:
                celda.number_format = FORMATO_ENTERO
                celda.alignment = Alignment(horizontal="right")
        if fila % 2 == 0:
            for col in range(1, 11):
                ws.cell(row=fila, column=col).fill = relleno_alterno
        fila += 1

    total_consumo_bn = sum((c.consumo_bn for c in resultado.consumo_por_equipo), 0)
    total_consumo_color = sum((c.consumo_color for c in resultado.consumo_por_equipo), 0)

    if not resultado.consumo_por_equipo:
        ws.cell(row=fila, column=1, value="Sin equipos con consumo registrado en este periodo.")
        fila += 1
    else:
        ws.cell(row=fila, column=1, value="Total consumo del periodo").font = Font(bold=True, color=AZUL_OSCURO)
        ws.cell(row=fila, column=9, value=total_consumo_bn).font = Font(bold=True, color=AZUL_OSCURO)
        ws.cell(row=fila, column=10, value=total_consumo_color).font = Font(bold=True, color=AZUL_OSCURO)
        for col in (9, 10):
            ws.cell(row=fila, column=col).number_format = FORMATO_ENTERO
            ws.cell(row=fila, column=col).alignment = Alignment(horizontal="right")
        borde_total = Border(
            top=Side(style="medium", color=AZUL_OSCURO),
            bottom=borde_fino,
            left=borde_fino,
            right=borde_fino,
        )
        for col in range(1, 11):
            ws.cell(row=fila, column=col).border = borde_total
        fila += 1

    fila += 2  # renglones en blanco antes del resumen
    monto_excedente_bn = Decimal(resultado.consumo_excedente_bn) * contrato.costo_excedente_bn
    monto_excedente_color = Decimal(resultado.consumo_excedente_color) * contrato.costo_excedente_color
    resumen = [
        ("Consumo total BN (copias)", total_consumo_bn, FORMATO_ENTERO),
        ("Copias incluidas BN", contrato.copias_incluidas_bn, FORMATO_ENTERO),
        (
            f"Excedente BN ({resultado.consumo_excedente_bn:,} copias × "
            f"{_tarifa(contrato.costo_excedente_bn, resultado.moneda)})",
            float(monto_excedente_bn),
            formato_moneda,
        ),
        ("Consumo total Color (copias)", total_consumo_color, FORMATO_ENTERO),
        ("Copias incluidas Color", contrato.copias_incluidas_color, FORMATO_ENTERO),
        (
            f"Excedente Color ({resultado.consumo_excedente_color:,} copias × "
            f"{_tarifa(contrato.costo_excedente_color, resultado.moneda)})",
            float(monto_excedente_color),
            formato_moneda,
        ),
        ("Monto renta", float(resultado.monto_renta), formato_moneda),
        (f"IVA ({contrato.iva_porcentaje}%)", float(resultado.monto_iva), formato_moneda),
        ("Total", float(resultado.monto_total), formato_moneda),
    ]
    ws.cell(row=fila, column=1, value="Resumen de facturación").font = Font(
        size=12, bold=True, color=AZUL_OSCURO
    )
    fila += 1
    for indice, (etiqueta, valor, formato) in enumerate(resumen):
        es_total = etiqueta == "Total"
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)
        celda_etiqueta = ws.cell(row=fila, column=1, value=etiqueta)
        celda_etiqueta.font = fuente_etiqueta
        celda_valor = ws.cell(row=fila, column=4, value=valor)
        celda_valor.number_format = formato
        celda_valor.alignment = Alignment(horizontal="right")

        if es_total:
            celda_etiqueta.font = Font(bold=True, size=12, color=AZUL_OSCURO)
            celda_valor.font = Font(bold=True, size=12, color=AZUL_OSCURO)
            borde_fila = Border(top=Side(style="medium", color=AZUL_OSCURO), bottom=borde_fino)
        else:
            borde_fila = borde_celda
            if indice % 2 == 1:
                for col in range(1, 5):
                    ws.cell(row=fila, column=col).fill = relleno_alterno
        for col in range(1, 5):
            ws.cell(row=fila, column=col).border = borde_fila
        fila += 1

    anchos = [18, 28, 16, 18, 17, 16, 14, 16, 12, 14]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
