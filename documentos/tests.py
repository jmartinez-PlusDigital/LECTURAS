import io
import shutil
import tempfile
from datetime import date
from decimal import Decimal

from django.core.files.base import ContentFile
from django.test import TestCase, override_settings
from openpyxl import load_workbook
from PIL import Image as PILImage

from core.models import Cliente, Contrato, EmpresaEmisora
from documentos.excel import generar_excel_factura
from documentos.pdf import _logo_data_uri, generar_pdf_factura
from facturacion.tipos import ConsumoEquipo, ResultadoFacturacion


def _png_de_prueba() -> bytes:
    buffer = io.BytesIO()
    PILImage.new("RGB", (12, 6), color="red").save(buffer, format="PNG")
    return buffer.getvalue()


class LogoEnDocumentosTestCase(TestCase):
    """El PDF y el Excel de factura usan el logo/nombre de `resultado.emisor`
    (ver core.models.EmpresaEmisora) en vez del EMPRESA_NOMBRE fijo de antes."""

    def setUp(self):
        self.cliente = Cliente.objects.create(nombre="Cliente Logo Test")
        self.contrato = Contrato.objects.create(
            numero_contrato="CT-LOGO-1",
            cliente=self.cliente,
            renta_base=Decimal("1000.00"),
            costo_excedente_bn=Decimal("0.50"),
            costo_excedente_color=Decimal("1.50"),
            dia_corte_facturacion=1,
            fecha_inicio=date(2026, 1, 1),
        )

    def _resultado(self, emisor) -> ResultadoFacturacion:
        return ResultadoFacturacion(
            contrato_id=self.contrato.id,
            periodo_mes=1,
            periodo_anio=2026,
            fecha_inicio=date(2026, 1, 1),
            fecha_fin=date(2026, 1, 31),
            moneda="MXN",
            emisor=emisor,
            consumo_por_equipo=[],
            consumo_excedente_bn=0,
            consumo_excedente_color=0,
            monto_renta=Decimal("1000.00"),
            monto_excedente=Decimal("0.00"),
            monto_iva=Decimal("160.00"),
            monto_total=Decimal("1160.00"),
        )

    def _emisor_con_logo(self, nombre: str) -> EmpresaEmisora:
        media_root = tempfile.mkdtemp()
        self.addCleanup(shutil.rmtree, media_root, ignore_errors=True)
        self.enterContext(override_settings(MEDIA_ROOT=media_root))
        emisor = EmpresaEmisora.objects.create(nombre=nombre)
        emisor.logo.save("logo.png", ContentFile(_png_de_prueba()), save=True)
        return emisor

    def test_pdf_sin_emisor_no_revienta(self):
        pdf_bytes = generar_pdf_factura(self._resultado(None), self.contrato)
        self.assertTrue(pdf_bytes.startswith(b"%PDF"))

    def test_excel_sin_emisor_no_revienta(self):
        excel_bytes = generar_excel_factura(self._resultado(None), self.contrato)
        self.assertGreater(len(excel_bytes), 0)

    def test_logo_data_uri_none_si_no_hay_emisor(self):
        self.assertIsNone(_logo_data_uri(None))

    def test_logo_data_uri_none_si_el_emisor_no_tiene_logo(self):
        emisor_sin_logo = EmpresaEmisora.objects.create(nombre="Sin Logo Test")
        self.assertIsNone(_logo_data_uri(emisor_sin_logo))

    def test_logo_data_uri_incrusta_el_logo_en_base64(self):
        emisor = self._emisor_con_logo("Con Logo PDF Test")

        data_uri = _logo_data_uri(emisor)

        self.assertTrue(data_uri.startswith("data:image/png;base64,"))

    def test_excel_incrusta_logo_cuando_el_emisor_lo_tiene(self):
        emisor = self._emisor_con_logo("Con Logo Excel Test")

        excel_bytes = generar_excel_factura(self._resultado(emisor), self.contrato)

        wb = load_workbook(io.BytesIO(excel_bytes))
        self.assertEqual(len(wb.active._images), 1)

    def test_excel_sin_logo_no_incrusta_ninguna_imagen(self):
        excel_bytes = generar_excel_factura(self._resultado(None), self.contrato)

        wb = load_workbook(io.BytesIO(excel_bytes))
        self.assertEqual(len(wb.active._images), 0)


class UbicacionEnDocumentosTestCase(TestCase):
    """Columna "Departamento / Ubicación" de la tabla de equipos (ver
    Asignacion.ubicacion, propagada vía ConsumoEquipo hasta el documento)."""

    def setUp(self):
        self.cliente = Cliente.objects.create(nombre="Cliente Ubicacion Test")
        self.contrato = Contrato.objects.create(
            numero_contrato="CT-UBIC-1",
            cliente=self.cliente,
            renta_base=Decimal("1000.00"),
            costo_excedente_bn=Decimal("0.50"),
            costo_excedente_color=Decimal("1.50"),
            dia_corte_facturacion=1,
            fecha_inicio=date(2026, 1, 1),
        )

    def _resultado_con_equipo(self, ubicacion: str) -> ResultadoFacturacion:
        consumo = ConsumoEquipo(
            asignacion_id=1,
            equipo_numero_serie="SN-UBIC-1",
            consumo_bn=100,
            consumo_color=20,
            lectura_anterior_bn=0,
            lectura_anterior_color=0,
            fecha_lectura_anterior=date(2026, 1, 1),
            lectura_actual_bn=100,
            lectura_actual_color=20,
            fecha_lectura_actual=date(2026, 1, 31),
            ubicacion=ubicacion,
        )
        return ResultadoFacturacion(
            contrato_id=self.contrato.id,
            periodo_mes=1,
            periodo_anio=2026,
            fecha_inicio=date(2026, 1, 1),
            fecha_fin=date(2026, 1, 31),
            moneda="MXN",
            emisor=None,
            consumo_por_equipo=[consumo],
            consumo_excedente_bn=0,
            consumo_excedente_color=0,
            monto_renta=Decimal("1000.00"),
            monto_excedente=Decimal("0.00"),
            monto_iva=Decimal("160.00"),
            monto_total=Decimal("1160.00"),
        )

    def test_excel_incluye_la_ubicacion_en_la_columna_correcta(self):
        resultado = self._resultado_con_equipo("Almacén")

        excel_bytes = generar_excel_factura(resultado, self.contrato)

        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active
        # fila_encabezado_tabla=7 (ver documentos/excel.py); primer equipo en
        # fila 8; columna 3 = Departamento / Ubicación.
        self.assertEqual(ws.cell(row=7, column=3).value, "Departamento / Ubicación")
        self.assertEqual(ws.cell(row=8, column=3).value, "Almacén")

    def test_pdf_con_ubicacion_no_revienta(self):
        resultado = self._resultado_con_equipo("Recepción")

        pdf_bytes = generar_pdf_factura(resultado, self.contrato)

        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
